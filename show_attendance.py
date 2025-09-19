import pandas as pd
from glob import glob
import os
import csv
import tkinter as tk
from tkinter import *
from tkinter import messagebox, filedialog, ttk
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def subjectchoose(text_to_speech):
    # Modern theme colors
    PRIMARY_COLOR = "#2563eb"
    SECONDARY_COLOR = "#3b82f6"
    ACCENT_COLOR = "#0ea5e9"
    BG_COLOR = "#e2e8f0"
    CARD_COLOR = "#f1f5f9"
    TEXT_COLOR = "#1e293b"
    SUCCESS_COLOR = "#16a34a"
    ERROR_COLOR = "#dc2626"
    ERROR_COLOR = "#ef4444"
    WARNING_COLOR = "#f59e0b"
    
    current_subject = ""
    current_df = None
    
    def safe_update_message(msg, color=None):
        try:
            if color:
                error_label.configure(text=msg, fg=color)
            else:
                error_label.configure(text=msg)
            subject.update()
        except:
            print(f"Message: {msg}")
    
    def get_available_subjects():
        """Get list of available subjects with attendance data"""
        subjects = []
        if os.path.exists("Attendance"):
            for item in os.listdir("Attendance"):
                item_path = os.path.join("Attendance", item)
                if os.path.isdir(item_path):
                    # Check if there are any CSV files in the subject folder
                    csv_files = glob(os.path.join(item_path, "*.csv"))
                    if csv_files:
                        subjects.append(item)
        return subjects
    
    def refresh_subjects():
        """Refresh the subject dropdown"""
        subjects = get_available_subjects()
        subject_combo['values'] = subjects
        if subjects:
            safe_update_message(f"✅ Found {len(subjects)} subjects with attendance data", SUCCESS_COLOR)
        else:
            safe_update_message("⚠️ No subjects found with attendance data", WARNING_COLOR)
    
    def calculate_attendance():
        global current_subject, current_df
        subject_name = subject_combo.get().strip()
        
        if not subject_name:
            safe_update_message("⚠️ Please select a subject!", ERROR_COLOR)
            text_to_speech("Please select a subject name.")
            return
        
        try:
            safe_update_message("📊 Processing attendance data...", WARNING_COLOR)
            
            # Find all CSV files for the subject
            filenames = glob(f"Attendance/{subject_name}/{subject_name}*.csv")
            
            if not filenames:
                safe_update_message(f"❌ No attendance files found for {subject_name}!", ERROR_COLOR)
                text_to_speech(f"No attendance files found for {subject_name}")
                return
            
            # Read and merge all CSV files
            df_list = []
            for file in filenames:
                try:
                    df = pd.read_csv(file)
                    if not df.empty:
                        df_list.append(df)
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue
            
            if not df_list:
                safe_update_message("❌ No valid attendance data found!", ERROR_COLOR)
                return
            
            # Merge all dataframes
            merged_df = df_list[0].copy()
            for df in df_list[1:]:
                merged_df = merged_df.merge(df, on=['Enrollment', 'Name'], how='outer', suffixes=('', '_y'))
                # Remove duplicate columns created by merge
                duplicate_cols = [col for col in merged_df.columns if col.endswith('_y')]
                for col in duplicate_cols:
                    original_col = col[:-2]  # Remove '_y' suffix
                    if original_col in merged_df.columns:
                        # Combine the values (take max since 1 means present, 0 means absent)
                        merged_df[original_col] = merged_df[[original_col, col]].max(axis=1)
                        # Drop the duplicate column
                        merged_df = merged_df.drop(columns=[col])
            
            # Fill missing values with 0
            merged_df = merged_df.fillna(0)
            
            # Calculate attendance percentage - exclude non-date columns
            date_columns = [col for col in merged_df.columns if col not in ['Enrollment', 'Name'] and not col.endswith('_y')]
            print(f"Debug: All columns: {list(merged_df.columns)}")
            print(f"Debug: Date columns identified: {date_columns}")
            
            if date_columns:
                # Convert attendance columns to numeric, replacing non-numeric values with 0
                for col in date_columns:
                    if col in merged_df.columns:  # Check if column exists
                        merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').fillna(0)
                
                # Filter out only valid date columns that exist in the dataframe
                valid_date_columns = [col for col in date_columns if col in merged_df.columns]
                
                if valid_date_columns:
                    try:
                        # Calculate attendance percentage
                        merged_df['Total_Classes'] = len(valid_date_columns)
                        # Ensure we're working with valid numeric data
                        numeric_cols = []
                        for col in valid_date_columns:
                            try:
                                merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').fillna(0)
                                numeric_cols.append(col)
                            except Exception as e:
                                print(f"Warning: Could not convert column {col} to numeric: {e}")
                        
                        if numeric_cols:
                            merged_df['Classes_Attended'] = merged_df[numeric_cols].sum(axis=1)
                            merged_df['Attendance_Percentage'] = (merged_df['Classes_Attended'] / merged_df['Total_Classes'] * 100).round(2)
                            
                            # Reorder columns
                            cols = ['Enrollment', 'Name'] + numeric_cols + ['Total_Classes', 'Classes_Attended', 'Attendance_Percentage']
                            merged_df = merged_df[cols]
                        else:
                            # No valid numeric columns
                            merged_df['Total_Classes'] = 0
                            merged_df['Classes_Attended'] = 0
                            merged_df['Attendance_Percentage'] = 0.0
                            
                    except Exception as e:
                        print(f"Error in attendance calculation: {e}")
                        # Fallback: basic data structure
                        merged_df['Total_Classes'] = 0
                        merged_df['Classes_Attended'] = 0
                        merged_df['Attendance_Percentage'] = 0.0
                else:
                    # If no valid date columns, just keep basic info
                    merged_df['Total_Classes'] = 0
                    merged_df['Classes_Attended'] = 0
                    merged_df['Attendance_Percentage'] = 0.0
            
            current_subject = subject_name
            current_df = merged_df
            
            # Save the processed data
            os.makedirs(f"Attendance/{subject_name}", exist_ok=True)
            output_file = f"Attendance/{subject_name}/attendance_summary.csv"
            merged_df.to_csv(output_file, index=False)
            
            # Display the data
            display_attendance_data(merged_df, subject_name)
            
            safe_update_message(f"✅ Attendance processed successfully for {subject_name}!", SUCCESS_COLOR)
            text_to_speech(f"Attendance calculated successfully for {subject_name}")
            
        except Exception as e:
            error_msg = f"❌ Error processing attendance: {str(e)}"
            safe_update_message(error_msg, ERROR_COLOR)
            text_to_speech(f"Error occurred: {str(e)}")
            print(f"Detailed error: {e}")
    
    def display_attendance_data(df, subject_name):
        """Display attendance data in a new window with better UI"""
        # Create new window
        display_window = Toplevel(subject)
        display_window.title(f"📊 {subject_name} - Attendance Report")
        display_window.geometry("1200x700")
        display_window.configure(bg=BG_COLOR)
        display_window.state('zoomed')  # Maximize window
        
        # Header frame
        header_frame = Frame(display_window, bg=BG_COLOR, height=80)
        header_frame.pack(fill=X, padx=20, pady=10)
        header_frame.pack_propagate(False)
        
        title_label = Label(
            header_frame,
            text=f"📊 Attendance Report - {subject_name}",
            bg=BG_COLOR,
            fg=ACCENT_COLOR,
            font=("Segoe UI", 24, "bold")
        )
        title_label.pack(pady=10)
        
        # Stats frame
        stats_frame = Frame(display_window, bg=BG_COLOR)
        stats_frame.pack(fill=X, padx=20, pady=5)
        
        total_students = len(df)
        avg_attendance = df['Attendance_Percentage'].mean() if 'Attendance_Percentage' in df.columns else 0
        total_classes = df['Total_Classes'].iloc[0] if 'Total_Classes' in df.columns and len(df) > 0 else 0
        
        stats_text = f"Total Students: {total_students} | Total Classes: {total_classes} | Average Attendance: {avg_attendance:.1f}%"
        stats_label = Label(
            stats_frame,
            text=stats_text,
            bg=BG_COLOR,
            fg=TEXT_COLOR,
            font=("Segoe UI", 14, "bold")
        )
        stats_label.pack()
        
        # Create frame for treeview with scrollbars
        tree_frame = Frame(display_window, bg=BG_COLOR)
        tree_frame.pack(expand=True, fill=BOTH, padx=20, pady=10)
        
        # Create treeview
        columns = list(df.columns)
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Configure column headings
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor=CENTER)
        
        # Add data to treeview
        for index, row in df.iterrows():
            values = []
            for col in columns:
                value = row[col]
                if col == 'Attendance_Percentage':
                    values.append(f"{value}%")
                else:
                    values.append(str(value))
            tree.insert('', END, values=values)
        
        # Add scrollbars
        v_scrollbar = Scrollbar(tree_frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=v_scrollbar.set)
        v_scrollbar.pack(side=RIGHT, fill=Y)
        
        h_scrollbar = Scrollbar(tree_frame, orient=HORIZONTAL, command=tree.xview)
        tree.configure(xscrollcommand=h_scrollbar.set)
        h_scrollbar.pack(side=BOTTOM, fill=X)
        
        tree.pack(expand=True, fill=BOTH)
        
        # Buttons frame
        buttons_frame = Frame(display_window, bg=BG_COLOR)
        buttons_frame.pack(fill=X, padx=20, pady=20)
        
        def download_excel():
            download_attendance_excel(df, subject_name, display_window)
        
        def download_csv():
            download_attendance_csv(df, subject_name, display_window)
        
        # Modern button style for display window
        def create_display_button(parent, text, command, color, emoji=""):
            btn = Button(
                parent,
                text=f"{emoji} {text}",
                command=command,
                bg=color,
                fg="white",
                font=("Segoe UI", 12, "bold"),
                relief=FLAT,
                bd=0,
                padx=20,
                pady=10,
                cursor="hand2"
            )
            
            def on_enter(e):
                btn.configure(relief=RAISED, bd=2)
            def on_leave(e):
                btn.configure(relief=FLAT, bd=0)
            
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
            return btn
        
        # Add buttons
        excel_btn = create_display_button(buttons_frame, "Download Excel", download_excel, SUCCESS_COLOR, "📊")
        excel_btn.pack(side=LEFT, padx=10)
        
        csv_btn = create_display_button(buttons_frame, "Download CSV", download_csv, PRIMARY_COLOR, "📄")
        csv_btn.pack(side=LEFT, padx=10)
        
        close_btn = create_display_button(buttons_frame, "Close", display_window.destroy, ERROR_COLOR, "❌")
        close_btn.pack(side=RIGHT, padx=10)
    
    def download_attendance_excel(df, subject_name, parent_window=None):
        """Download attendance data as Excel file"""
        try:
            print(f"Debug: Starting Excel download for subject: {subject_name}")
            print(f"Debug: DataFrame shape: {df.shape}")
            print(f"Debug: DataFrame columns: {list(df.columns)}")
            
            # Ask user for save location
            # Ask user for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Attendance Report as Excel",
                initialfile=f"{subject_name}_attendance_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                print("Debug: No filename selected for Excel download")
                return
            
            print(f"Debug: Saving Excel to: {filename}")
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = f"{subject_name} Attendance"
            
            # Write header information
            ws['A1'] = f"{subject_name} - Attendance Report"
            ws['A2'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Total Students: {len(df)}"
            
            if 'Attendance_Percentage' in df.columns:
                ws['A4'] = f"Average Attendance: {df['Attendance_Percentage'].mean():.2f}%"
            
            # Style header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=14)
            
            for row in range(1, 5):
                cell = ws[f'A{row}']
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data starting from row 6
            start_row = 6
            
            # Write column headers
            for col, column_name in enumerate(df.columns, 1):
                cell = ws.cell(row=start_row, column=col, value=column_name)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_idx, (index, row) in enumerate(df.iterrows(), start_row + 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center")
                    
                    # Color code attendance percentage
                    if df.columns[col_idx - 1] == 'Attendance_Percentage':
                        if value >= 75:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value >= 50:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(df), 
                                   min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            print("Debug: Saving Excel workbook...")
            wb.save(filename)
            print("Debug: Excel workbook saved successfully")
            
            # Show success message
            if parent_window:
                messagebox.showinfo("Success", f"Excel file saved successfully!\nLocation: {filename}", parent=parent_window)
            else:
                try:
                    safe_update_message(f"✅ Excel file saved successfully: {filename}", SUCCESS_COLOR)
                except:
                    messagebox.showinfo("Success", f"Excel file saved successfully!\nLocation: {filename}")
            
            text_to_speech("Excel file downloaded successfully")
            
        except Exception as e:
            error_msg = f"❌ Error saving Excel file: {str(e)}"
            print(f"Debug Excel Error: {error_msg}")
            import traceback
            print(f"Debug Excel Traceback: {traceback.format_exc()}")
            
            if parent_window:
                messagebox.showerror("Error", error_msg, parent=parent_window)
            else:
                try:
                    safe_update_message(error_msg, ERROR_COLOR)
                except:
                    messagebox.showerror("Error", error_msg)
            
            text_to_speech("Error saving Excel file")
    
    def download_attendance_csv(df, subject_name, parent_window=None):
        """Download attendance data as CSV file"""
        try:
            print(f"Debug: Starting CSV download for subject: {subject_name}")
            print(f"Debug: DataFrame shape: {df.shape}")
            print(f"Debug: DataFrame columns: {list(df.columns)}")
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Save Attendance Report as CSV",
                initialfile=f"{subject_name}_attendance_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                print(f"Debug: Saving CSV to: {filename}")
                df.to_csv(filename, index=False)
                print("Debug: CSV file saved successfully")
                
                # Show success message
                if parent_window:
                    messagebox.showinfo("Success", f"CSV file saved successfully!\nLocation: {filename}", parent=parent_window)
                else:
                    try:
                        safe_update_message(f"✅ CSV file saved successfully: {filename}", SUCCESS_COLOR)
                    except:
                        messagebox.showinfo("Success", f"CSV file saved successfully!\nLocation: {filename}")
                
                text_to_speech("CSV file downloaded successfully")
            else:
                print("Debug: No filename selected for CSV download")
                
        except Exception as e:
            error_msg = f"❌ Error saving CSV file: {str(e)}"
            print(f"Debug CSV Error: {error_msg}")
            import traceback
            print(f"Debug CSV Traceback: {traceback.format_exc()}")
            
            if parent_window:
                messagebox.showerror("Error", error_msg, parent=parent_window)
            else:
                try:
                    safe_update_message(error_msg, ERROR_COLOR)
                except:
                    messagebox.showerror("Error", error_msg)
            
            text_to_speech("Error saving CSV file")
    
    def open_folder():
        """Open the attendance folder"""
        subject_name = subject_combo.get().strip()
        if not subject_name:
            safe_update_message("⚠️ Please select a subject!", ERROR_COLOR)
            return
        
        try:
            folder_path = f"Attendance/{subject_name}"
            if os.path.exists(folder_path):
                os.startfile(folder_path)
                safe_update_message("📁 Opening attendance folder...", SUCCESS_COLOR)
            else:
                safe_update_message("❌ Subject folder not found!", ERROR_COLOR)
        except Exception as e:
            safe_update_message(f"❌ Error opening folder: {str(e)}", ERROR_COLOR)
    
    # Create main window (using Tk since this might be called from main app)
    try:
        subject = Toplevel()
    except:
        subject = Tk()
    subject.title("📊 View Attendance Reports")
    subject.geometry("700x600")
    subject.configure(bg=BG_COLOR)
    subject.resizable(True, True)
    
    # Main container
    main_frame = Frame(subject, bg=BG_COLOR)
    main_frame.pack(expand=True, fill=BOTH, padx=40, pady=30)
    
    # Header Section
    title_label = Label(
        main_frame,
        text="📊 View Attendance Reports",
        bg=BG_COLOR,
        fg=ACCENT_COLOR,
        font=("Segoe UI", 28, "bold")
    )
    title_label.pack(pady=(0, 10))
    
    subtitle = Label(
        main_frame,
        text="Generate detailed attendance reports with Excel download",
        bg=BG_COLOR,
        fg=TEXT_COLOR,
        font=("Segoe UI", 16)
    )
    subtitle.pack(pady=(0, 30))
    
    # Input Section
    input_card = Frame(main_frame, bg=CARD_COLOR, relief=FLAT, bd=0)
    input_card.pack(fill=X, pady=(0, 20))
    
    input_inner = Frame(input_card, bg=CARD_COLOR)
    input_inner.pack(pady=30, padx=30, fill=X)
    
    sub_label = Label(
        input_inner,
        text="Select Subject:",
        bg=CARD_COLOR,
        fg=TEXT_COLOR,
        font=("Segoe UI", 14, "bold")
    )
    sub_label.pack(pady=(0, 10), anchor=W)
    
    # Subject selection with combobox
    combo_frame = Frame(input_inner, bg=CARD_COLOR)
    combo_frame.pack(fill=X, pady=(0, 10))
    
    subject_combo = ttk.Combobox(
        combo_frame,
        font=("Segoe UI", 14),
        state="readonly",
        width=30
    )
    subject_combo.pack(side=LEFT, padx=(0, 10), ipady=8, fill=X, expand=True)
    
    refresh_btn = Button(
        combo_frame,
        text="🔄 Refresh",
        command=refresh_subjects,
        bg=SECONDARY_COLOR,
        fg="white",
        font=("Segoe UI", 12, "bold"),
        relief=FLAT,
        bd=0,
        padx=15,
        pady=8,
        cursor="hand2"
    )
    refresh_btn.pack(side=RIGHT)
    
    # Error/Success message area
    error_label = Label(
        main_frame,
        text="",
        bg=BG_COLOR,
        fg=SUCCESS_COLOR,
        font=("Segoe UI", 12),
        wraplength=600
    )
    error_label.pack(pady=(10, 20))
    
    # Modern Button Styles
    def modern_button(parent, text, command, color, emoji=""):
        btn = Button(
            parent,
            text=f"{emoji} {text}",
            command=command,
            bg=color,
            fg="white",
            font=("Segoe UI", 14, "bold"),
            relief=FLAT,
            bd=0,
            padx=30,
            pady=15,
            cursor="hand2"
        )
        btn.pack(pady=8, fill=X)
        
        # Hover effects
        def on_enter(e):
            btn.configure(relief=RAISED, bd=2)
        def on_leave(e):
            btn.configure(relief=FLAT, bd=0)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn
    
    # Buttons Section
    buttons_frame = Frame(main_frame, bg=BG_COLOR)
    buttons_frame.pack(fill=X, pady=(20, 0))
    
    view_btn = modern_button(buttons_frame, "Generate Report", calculate_attendance, PRIMARY_COLOR, "📊")
    folder_btn = modern_button(buttons_frame, "Open Folder", open_folder, SECONDARY_COLOR, "📁")
    
    # Initialize with available subjects
    refresh_subjects()
    
    subject.mainloop()


def subjectchoose_with_subject(text_to_speech, preselected_subject):
    """Show attendance reports with a pre-selected subject"""
    # Modern theme colors
    PRIMARY_COLOR = "#2563eb"
    SECONDARY_COLOR = "#7c3aed"
    ACCENT_COLOR = "#0ea5e9"
    BG_COLOR = "#e2e8f0"
    CARD_COLOR = "#f1f5f9"
    TEXT_COLOR = "#1e293b"
    SUCCESS_COLOR = "#16a34a"
    ERROR_COLOR = "#ef4444"
    WARNING_COLOR = "#f59e0b"
    
    current_subject = preselected_subject
    current_df = None
    
    def safe_update_message(msg, color=None):
        try:
            if color:
                error_label.configure(text=msg, fg=color)
            else:
                error_label.configure(text=msg)
            subject.update()
        except:
            print(f"Message: {msg}")
    
    def get_available_subjects():
        """Get list of available subjects with attendance data"""
        subjects = []
        if os.path.exists("Attendance"):
            for item in os.listdir("Attendance"):
                item_path = os.path.join("Attendance", item)
                if os.path.isdir(item_path):
                    # Check if there are any CSV files in the subject folder
                    csv_files = glob(os.path.join(item_path, "*.csv"))
                    if csv_files:
                        subjects.append(item)
        return subjects
    
    def load_attendance_data(subject_name):
        """Load attendance data for a given subject"""
        try:
            subject_path = os.path.join("Attendance", subject_name)
            if not os.path.exists(subject_path):
                return None, f"No attendance data found for subject: {subject_name}"
            
            # Get all CSV files for this subject (excluding summary files)
            csv_files = glob(os.path.join(subject_path, f"{subject_name}_*.csv"))
            
            if not csv_files:
                return None, f"No attendance records found for subject: {subject_name}"
            
            # Combine all attendance files
            all_attendance = []
            for file in csv_files:
                try:
                    df = pd.read_csv(file)
                    # Extract date from filename
                    filename = os.path.basename(file)
                    # Format: subject_YYYY-MM-DD_HH-MM-SS.csv
                    date_part = filename.split('_')[1]  # Get YYYY-MM-DD part
                    df['Date'] = date_part
                    all_attendance.append(df)
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue
            
            if not all_attendance:
                return None, f"Could not read any attendance files for {subject_name}"
            
            # Combine all dataframes
            combined_df = pd.concat(all_attendance, ignore_index=True)
            return combined_df, None
            
        except Exception as e:
            return None, f"Error loading attendance data: {str(e)}"
    
    def calculate_attendance():
        """Calculate and display attendance statistics for the preselected subject"""
        nonlocal current_df, current_subject
        
        if not current_subject:
            safe_update_message("No subject selected", ERROR_COLOR)
            return
        
        safe_update_message("Loading attendance data...", WARNING_COLOR)
        
        # Load attendance data
        df, error = load_attendance_data(current_subject)
        
        if error:
            safe_update_message(error, ERROR_COLOR)
            text_to_speech("Error loading attendance data")
            return
        
        current_df = df
        
        # Calculate statistics
        total_days = len(df['Date'].unique()) if not df.empty else 0
        total_students = len(df['Enrollment'].unique()) if not df.empty else 0
        total_present = len(df) if not df.empty else 0
        
        # Create summary statistics
        if not df.empty:
            attendance_summary = df.groupby('Enrollment').agg({
                'Name': 'first',
                'Date': 'count'
            }).rename(columns={'Date': 'Days_Present'})
            
            attendance_summary['Total_Days'] = total_days
            attendance_summary['Attendance_Percentage'] = (attendance_summary['Days_Present'] / total_days * 100).round(2)
            attendance_summary = attendance_summary.reset_index()
            
            # Display results in a new window
            display_attendance_report(attendance_summary, current_subject, total_days)
            safe_update_message(f"✅ Report generated for {current_subject} - {total_students} students, {total_days} days", SUCCESS_COLOR)
            text_to_speech(f"Attendance report generated for {current_subject}")
        else:
            safe_update_message("No attendance data available", ERROR_COLOR)
            text_to_speech("No attendance data found")
    
    def display_attendance_report(summary_df, subject_name, total_days):
        """Display attendance report in a new window with download option"""
        report_window = Toplevel()
        report_window.title(f"📊 Attendance Report - {subject_name}")
        report_window.geometry("900x600")
        report_window.configure(bg=BG_COLOR)
        
        # Header
        header_frame = Frame(report_window, bg=BG_COLOR)
        header_frame.pack(fill=X, padx=20, pady=20)
        
        title_label = Label(
            header_frame,
            text=f"📊 Attendance Report - {subject_name}",
            bg=BG_COLOR,
            fg=ACCENT_COLOR,
            font=("Segoe UI", 20, "bold")
        )
        title_label.pack()
        
        stats_label = Label(
            header_frame,
            text=f"Total Days: {total_days} | Total Students: {len(summary_df)}",
            bg=BG_COLOR,
            fg=TEXT_COLOR,
            font=("Segoe UI", 12)
        )
        stats_label.pack(pady=(5, 0))
        
        # Create treeview for data display
        tree_frame = Frame(report_window, bg=BG_COLOR)
        tree_frame.pack(expand=True, fill=BOTH, padx=20, pady=(0, 20))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=HORIZONTAL)
        
        # Treeview
        tree = ttk.Treeview(
            tree_frame,
            columns=('Enrollment', 'Name', 'Days_Present', 'Total_Days', 'Attendance_Percentage'),
            show='headings',
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )
        
        # Configure scrollbars
        v_scrollbar.configure(command=tree.yview)
        h_scrollbar.configure(command=tree.xview)
        
        # Column headings
        tree.heading('Enrollment', text='Enrollment')
        tree.heading('Name', text='Student Name')
        tree.heading('Days_Present', text='Days Present')
        tree.heading('Total_Days', text='Total Days')
        tree.heading('Attendance_Percentage', text='Attendance %')
        
        # Column widths
        tree.column('Enrollment', width=100)
        tree.column('Name', width=200)
        tree.column('Days_Present', width=100)
        tree.column('Total_Days', width=100)
        tree.column('Attendance_Percentage', width=120)
        
        # Insert data
        for _, row in summary_df.iterrows():
            percentage = row['Attendance_Percentage']
            # Color code based on attendance percentage
            tags = ()
            if percentage >= 75:
                tags = ('good',)
            elif percentage >= 60:
                tags = ('average',)
            else:
                tags = ('poor',)
            
            tree.insert('', 'end', values=(
                row['Enrollment'],
                row['Name'],
                row['Days_Present'],
                row['Total_Days'],
                f"{percentage}%"
            ), tags=tags)
        
        # Configure row colors
        tree.tag_configure('good', background='#d4edda')
        tree.tag_configure('average', background='#fff3cd')
        tree.tag_configure('poor', background='#f8d7da')
        
        # Pack treeview and scrollbars
        tree.pack(side=LEFT, expand=True, fill=BOTH)
        v_scrollbar.pack(side=RIGHT, fill=Y)
        h_scrollbar.pack(side=BOTTOM, fill=X)
        
        # Download button
        download_frame = Frame(report_window, bg=BG_COLOR)
        download_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        def download_report():
            download_attendance_csv(summary_df, subject_name, report_window)
        
        download_btn = Button(
            download_frame,
            text="📥 Download Excel Report",
            command=download_report,
            bg=SUCCESS_COLOR,
            fg="white",
            font=("Segoe UI", 12, "bold"),
            relief=FLAT,
            bd=0,
            padx=20,
            pady=10,
            cursor="hand2"
        )
        download_btn.pack()
    
    def download_attendance_csv(df, subject_name, parent_window=None):
        """Download attendance data as CSV file"""
        try:
            # Ask user where to save
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Save Attendance Report as CSV",
                initialfile=f"{subject_name}_attendance_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                print(f"Debug: Saving CSV to: {filename}")
                # Save as CSV file
                df.to_csv(filename, index=False)
                
                if parent_window:
                    messagebox.showinfo("Success", f"CSV file saved successfully!\nLocation: {filename}", parent=parent_window)
                else:
                    messagebox.showinfo("Success", f"CSV file saved successfully!\nLocation: {filename}")
                
                text_to_speech("CSV report downloaded successfully")
            else:
                print("Debug: No filename selected for CSV download")
        
        except Exception as e:
            error_msg = f"Error saving CSV file: {str(e)}"
            print(f"Debug CSV Error: {error_msg}")
            if parent_window:
                messagebox.showerror("Error", error_msg, parent=parent_window)
            else:
                messagebox.showerror("Error", error_msg)
            text_to_speech("Error saving CSV file")
    
    def open_folder():
        """Open the attendance folder"""
        try:
            import subprocess
            import platform
            
            attendance_folder = os.path.abspath("Attendance")
            if os.path.exists(attendance_folder):
                if platform.system() == "Windows":
                    subprocess.Popen(['explorer', attendance_folder])
                elif platform.system() == "Darwin":  # macOS
                    subprocess.Popen(['open', attendance_folder])
                else:  # Linux
                    subprocess.Popen(['xdg-open', attendance_folder])
                text_to_speech("Opening attendance folder")
            else:
                messagebox.showerror("Error", "Attendance folder not found!")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder: {str(e)}")
    
    # Create main window (using Toplevel since this is called from another window)
    subject = Toplevel()
    subject.title(f"📊 Attendance Report - {preselected_subject}")
    subject.geometry("700x500")
    subject.configure(bg=BG_COLOR)
    subject.resizable(True, True)
    
    # Main container
    main_frame = Frame(subject, bg=BG_COLOR)
    main_frame.pack(expand=True, fill=BOTH, padx=40, pady=30)
    
    # Header Section
    title_label = Label(
        main_frame,
        text=f"📊 Attendance Report - {preselected_subject}",
        bg=BG_COLOR,
        fg=ACCENT_COLOR,
        font=("Segoe UI", 24, "bold")
    )
    title_label.pack(pady=(0, 10))
    
    subtitle = Label(
        main_frame,
        text="Generate detailed attendance reports and statistics",
        bg=BG_COLOR,
        fg=TEXT_COLOR,
        font=("Segoe UI", 12)
    )
    subtitle.pack(pady=(0, 30))
    
    # Subject info
    subject_info = Label(
        main_frame,
        text=f"Subject: {preselected_subject}",
        bg=BG_COLOR,
        fg=ACCENT_COLOR,
        font=("Segoe UI", 16, "bold")
    )
    subject_info.pack(pady=(0, 20))
    
    # Error/Success message area
    error_label = Label(
        main_frame,
        text="Click 'Generate Report' to view attendance statistics",
        bg=BG_COLOR,
        fg=SUCCESS_COLOR,
        font=("Segoe UI", 12),
        wraplength=600
    )
    error_label.pack(pady=(10, 20))
    
    # Modern Button Styles
    def modern_button(parent, text, command, color, emoji=""):
        btn = Button(
            parent,
            text=f"{emoji} {text}",
            command=command,
            bg=color,
            fg="white",
            font=("Segoe UI", 14, "bold"),
            relief=FLAT,
            bd=0,
            padx=30,
            pady=15,
            cursor="hand2"
        )
        btn.pack(pady=8, fill=X)
        
        # Hover effects
        def on_enter(e):
            btn.configure(relief=RAISED, bd=2)
        def on_leave(e):
            btn.configure(relief=FLAT, bd=0)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn
    
    # Buttons Section
    buttons_frame = Frame(main_frame, bg=BG_COLOR)
    buttons_frame.pack(fill=X, pady=(20, 0))
    
    view_btn = modern_button(buttons_frame, "Generate Report", calculate_attendance, PRIMARY_COLOR, "📊")
    folder_btn = modern_button(buttons_frame, "Open Attendance Folder", open_folder, SECONDARY_COLOR, "📁")
    
    # Auto-generate report for the preselected subject
    subject.after(100, calculate_attendance)  # Auto-run after window loads
    
    subject.mainloop()
