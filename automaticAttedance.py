import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox
import os, cv2
import csv
import numpy as np
from PIL import ImageTk, Image
import pandas as pd
import datetime
import time
import tkinter.font as font
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

haarcasecade_path = "haarcascade_frontalface_default.xml"
trainimagelabel_path = (
    "TrainingImageLabel\\Trainner.yml"
)
trainimage_path = "TrainingImage"
studentdetail_path = (
    "StudentDetails\\studentdetails.csv"
)
attendance_path = "Attendance"

def download_excel_file(attendance_df, subject_name, date):
    """Download attendance data as formatted Excel file"""
    try:
        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialname=f"{subject_name}_attendance_{date}.xlsx",
            title="Save Attendance Excel File"
        )
        
        if not filename:
            return
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{subject_name} Attendance"
        
        # Write header information
        ws['A1'] = f"{subject_name} - Attendance Report"
        ws['A2'] = f"Date: {date}"
        ws['A3'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Students Present: {len(attendance_df)}"
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        
        for row in range(1, 5):
            cell = ws[f'A{row}']
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data starting from row 6
        start_row = 6
        
        # Write column headers
        headers = ['Enrollment', 'Name', 'Date', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Write attendance data
        for row_idx, (index, row) in enumerate(attendance_df.iterrows(), start_row + 1):
            ws.cell(row=row_idx, column=1, value=int(row['Enrollment']))
            ws.cell(row=row_idx, column=2, value=str(row['Name']).strip('[]'))
            ws.cell(row=row_idx, column=3, value=date)
            ws.cell(row=row_idx, column=4, value="Present")
            
            # Center align data
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
                
            # Color code present students with green
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(attendance_df), 
                               min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        column_widths = [15, 25, 12, 10]  # Custom widths for each column
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Save the file
        wb.save(filename)
        
        # Show success message
        messagebox.showinfo(
            "Excel Downloaded",
            f"✅ Excel file saved successfully!\n\nLocation: {filename}\n\nStudents recorded: {len(attendance_df)}",
            icon='info'
        )
        
    except Exception as e:
        messagebox.showerror(
            "Download Error",
            f"❌ Error saving Excel file: {str(e)}",
            icon='error'
        )

# for choose subject and fill attendance
def subjectChoose(text_to_speech):
    def FillAttendance():
        sub = tx.get()
        now = time.time()
        future = now + 20
        print(now)
        print(future)
        if sub == "":
            t = "Please enter the subject name!!!"
            text_to_speech(t)
        else:
            try:
                recognizer = cv2.face.LBPHFaceRecognizer_create()
                try:
                    recognizer.read(trainimagelabel_path)
                except:
                    e = "Model not found,please train model"
                    Notifica.configure(
                        text=e,
                        bg="black",
                        fg="yellow",
                        width=33,
                        font=("times", 15, "bold"),
                    )
                    Notifica.place(x=20, y=250)
                    text_to_speech(e)
                facecasCade = cv2.CascadeClassifier(haarcasecade_path)
                df = pd.read_csv(studentdetail_path)
                cam = cv2.VideoCapture(0)
                font = cv2.FONT_HERSHEY_SIMPLEX
                col_names = ["Enrollment", "Name"]
                attendance = pd.DataFrame(columns=col_names)
                while True:
                    ___, im = cam.read()
                    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
                    faces = facecasCade.detectMultiScale(gray, 1.2, 5)
                    for (x, y, w, h) in faces:
                        global Id

                        Id, conf = recognizer.predict(gray[y : y + h, x : x + w])
                        if conf < 70:
                            print(conf)
                            global Subject
                            global aa
                            global date
                            global timeStamp
                            Subject = tx.get()
                            ts = time.time()
                            date = datetime.datetime.fromtimestamp(ts).strftime(
                                "%Y-%m-%d"
                            )
                            timeStamp = datetime.datetime.fromtimestamp(ts).strftime(
                                "%H:%M:%S"
                            )
                            aa = df.loc[df["Enrollment"] == Id]["Name"].values
                            global tt
                            tt = str(Id) + "-" + aa
                            # En='1604501160'+str(Id)
                            attendance.loc[len(attendance)] = [
                                Id,
                                aa,
                            ]
                            cv2.rectangle(im, (x, y), (x + w, y + h), (0, 260, 0), 4)
                            cv2.putText(
                                im, str(tt), (x + h, y), font, 1, (255, 255, 0,), 4
                            )
                        else:
                            Id = "Unknown"
                            tt = str(Id)
                            cv2.rectangle(im, (x, y), (x + w, y + h), (0, 25, 255), 7)
                            cv2.putText(
                                im, str(tt), (x + h, y), font, 1, (0, 25, 255), 4
                            )
                    if time.time() > future:
                        break

                    attendance = attendance.drop_duplicates(
                        ["Enrollment"], keep="first"
                    )
                    cv2.imshow("Filling Attendance...", im)
                    key = cv2.waitKey(30) & 0xFF
                    if key == 27:
                        break

                ts = time.time()
                print(aa)
                # attendance["date"] = date
                # attendance["Attendance"] = "P"
                attendance[date] = 1
                date = datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
                timeStamp = datetime.datetime.fromtimestamp(ts).strftime("%H:%M:%S")
                Hour, Minute, Second = timeStamp.split(":")
                # fileName = "Attendance/" + Subject + ".csv"
                path = os.path.join(attendance_path, Subject)
                if not os.path.exists(path):
                    os.makedirs(path)
                fileName = (
                    f"{path}/"
                    + Subject
                    + "_"
                    + date
                    + "_"
                    + Hour
                    + "-"
                    + Minute
                    + "-"
                    + Second
                    + ".csv"
                )
                attendance = attendance.drop_duplicates(["Enrollment"], keep="first")
                print(attendance)
                attendance.to_csv(fileName, index=False)

                m = "Attendance Filled Successfully of " + Subject
                Notifica.configure(
                    text=m,
                    bg="black",
                    fg="yellow",
                    width=33,
                    relief=RIDGE,
                    bd=5,
                    font=("times", 15, "bold"),
                )
                text_to_speech(m)

                Notifica.place(x=20, y=250)

                cam.release()
                cv2.destroyAllWindows()

                import csv
                import tkinter

                root = tkinter.Tk()
                root.title("Attendance of " + Subject)
                root.configure(background="black")
                cs = os.path.join(path, fileName)
                print(cs)
                with open(cs, newline="") as file:
                    reader = csv.reader(file)
                    r = 0

                    for col in reader:
                        c = 0
                        for row in col:

                            label = tkinter.Label(
                                root,
                                width=10,
                                height=1,
                                fg="yellow",
                                font=("times", 15, " bold "),
                                bg="black",
                                text=row,
                                relief=tkinter.RIDGE,
                            )
                            label.grid(row=r, column=c)
                            c += 1
                        r += 1
                root.mainloop()
                print(attendance)
                
                # Show attendance completion options
                def show_attendance_options():
                    """Show comprehensive attendance management interface like the second image"""
                    options_window = tkinter.Toplevel()
                    options_window.title("CLASS VISION - Attendance Management")
                    options_window.geometry("1000x700")
                    options_window.configure(bg="#1a1a2e")
                    options_window.resizable(True, True)
                    
                    # Center the window
                    options_window.transient(root)
                    options_window.grab_set()
                    
                    # Color scheme matching the second image
                    DARK_BG = "#1a1a2e"
                    CARD_BG = "#16213e"
                    ACCENT_COLOR = "#0F3460"
                    CYAN_COLOR = "#00d4ff"
                    BUTTON_COLOR = "#00a8cc"
                    TEXT_COLOR = "#ffffff"
                    
                    # Header with logo and title
                    header_frame = tkinter.Frame(options_window, bg=DARK_BG, height=100)
                    header_frame.pack(fill="x", padx=20, pady=(20, 10))
                    header_frame.pack_propagate(False)
                    
                    # Try to load logo
                    try:
                        from PIL import Image, ImageTk
                        logo_img = Image.open("AMS.ico").resize((50, 50), Image.Resampling.LANCZOS)
                        logo_photo = ImageTk.PhotoImage(logo_img)
                        logo_label = tkinter.Label(header_frame, image=logo_photo, bg=DARK_BG)
                        logo_label.image = logo_photo  # Keep reference
                        logo_label.pack(side="left", padx=(20, 15))
                    except:
                        # Fallback text logo
                        logo_label = tkinter.Label(
                            header_frame, 
                            text="✓", 
                            bg=DARK_BG, 
                            fg=CYAN_COLOR, 
                            font=("Segoe UI", 30, "bold")
                        )
                        logo_label.pack(side="left", padx=(20, 15))
                    
                    title_frame = tkinter.Frame(header_frame, bg=DARK_BG)
                    title_frame.pack(side="left", fill="both", expand=True)
                    
                    main_title = tkinter.Label(
                        title_frame,
                        text="CLASS VISION",
                        bg=DARK_BG,
                        fg=CYAN_COLOR,
                        font=("Segoe UI", 24, "bold")
                    )
                    main_title.pack(anchor="w")
                    
                    subtitle = tkinter.Label(
                        title_frame,
                        text="AI-Powered Face Recognition Attendance System",
                        bg=DARK_BG,
                        fg="#8892b0",
                        font=("Segoe UI", 12)
                    )
                    subtitle.pack(anchor="w")
                    
                    # Divider line
                    divider = tkinter.Frame(options_window, bg=CYAN_COLOR, height=3)
                    divider.pack(fill="x", padx=20, pady=(0, 20))
                    
                    # Welcome message
                    welcome_frame = tkinter.Frame(options_window, bg=CARD_BG, relief="solid", bd=1)
                    welcome_frame.pack(fill="x", padx=50, pady=(10, 30))
                    
                    welcome_title = tkinter.Label(
                        welcome_frame,
                        text=f"Attendance Recorded Successfully for {Subject}!",
                        bg=CARD_BG,
                        fg=CYAN_COLOR,
                        font=("Segoe UI", 20, "bold")
                    )
                    welcome_title.pack(pady=20)
                    
                    # Main content area with cards
                    content_frame = tkinter.Frame(options_window, bg=DARK_BG)
                    content_frame.pack(expand=True, fill="both", padx=50, pady=20)
                    
                    def create_action_card(parent, title, description, command, icon=""):
                        # Card container
                        card_frame = tkinter.Frame(parent, bg=CARD_BG, relief="solid", bd=1)
                        card_frame.pack(side="left", fill="both", expand=True, padx=15)
                        
                        # Card content
                        card_content = tkinter.Frame(card_frame, bg=CARD_BG)
                        card_content.pack(fill="both", expand=True, padx=20, pady=20)
                        
                        # Icon
                        icon_label = tkinter.Label(
                            card_content,
                            text=icon,
                            bg=CARD_BG,
                            fg=TEXT_COLOR,
                            font=("Segoe UI", 40)
                        )
                        icon_label.pack(pady=(10, 20))
                        
                        # Title
                        title_label = tkinter.Label(
                            card_content,
                            text=title,
                            bg=CARD_BG,
                            fg=TEXT_COLOR,
                            font=("Segoe UI", 18, "bold")
                        )
                        title_label.pack(pady=(0, 10))
                        
                        # Description
                        desc_label = tkinter.Label(
                            card_content,
                            text=description,
                            bg=CARD_BG,
                            fg="#8892b0",
                            font=("Segoe UI", 11),
                            wraplength=250,
                            justify="center"
                        )
                        desc_label.pack(pady=(0, 20))
                        
                        # Action button
                        action_btn = tkinter.Button(
                            card_content,
                            text=f"Open {title}",
                            command=command,
                            bg=BUTTON_COLOR,
                            fg="white",
                            font=("Segoe UI", 12, "bold"),
                            relief="flat",
                            bd=0,
                            padx=25,
                            pady=12,
                            cursor="hand2"
                        )
                        action_btn.pack()
                        
                        # Hover effects
                        def on_enter(e):
                            action_btn.configure(bg="#0097b8")
                            card_frame.configure(bg="#1e2951")
                            card_content.configure(bg="#1e2951")
                            icon_label.configure(bg="#1e2951")
                            title_label.configure(bg="#1e2951")
                            desc_label.configure(bg="#1e2951")
                        
                        def on_leave(e):
                            action_btn.configure(bg=BUTTON_COLOR)
                            card_frame.configure(bg=CARD_BG)
                            card_content.configure(bg=CARD_BG)
                            icon_label.configure(bg=CARD_BG)
                            title_label.configure(bg=CARD_BG)
                            desc_label.configure(bg=CARD_BG)
                        
                        card_frame.bind("<Enter>", on_enter)
                        card_frame.bind("<Leave>", on_leave)
                        action_btn.bind("<Enter>", on_enter)
                        action_btn.bind("<Leave>", on_leave)
                        
                        return card_frame
                    
                    def download_excel():
                        """Download Excel file"""
                        options_window.destroy()
                        download_excel_file(attendance, Subject, date)
                        text_to_speech("Excel file downloaded successfully")
                    
                    def view_attendance():
                        """Open attendance viewer"""
                        options_window.destroy()
                        root.destroy()  # Close the attendance display
                        import show_attendance
                        show_attendance.subjectchoose_with_subject(text_to_speech, Subject)
                    
                    def view_all_attendance():
                        """Open general attendance viewer"""
                        options_window.destroy()
                        root.destroy()  # Close the attendance display
                        import show_attendance
                        show_attendance.subjectchoose(text_to_speech)
                    
                    # Create action cards
                    create_action_card(
                        content_frame,
                        "Download Excel",
                        "Download the attendance data as a formatted Excel spreadsheet with statistics and charts.",
                        download_excel,
                        "📥"
                    )
                    
                    create_action_card(
                        content_frame,
                        "View Attendance",
                        f"Browse and analyze attendance records for {Subject}, generate reports, and view attendance statistics.",
                        view_attendance,
                        "📊"
                    )
                    
                    create_action_card(
                        content_frame,
                        "All Subjects",
                        "Browse and analyze attendance records for all subjects, generate comprehensive reports.",
                        view_all_attendance,
                        "📋"
                    )
                    
                    # Bottom section with exit button
                    bottom_frame = tkinter.Frame(options_window, bg=DARK_BG, height=80)
                    bottom_frame.pack(fill="x", side="bottom", padx=50, pady=20)
                    bottom_frame.pack_propagate(False)
                    
                    def close_window():
                        options_window.destroy()
                        text_to_speech("Attendance management closed")
                    
                    exit_btn = tkinter.Button(
                        bottom_frame,
                        text="✖️ Close",
                        command=close_window,
                        bg="#ef4444",
                        fg="white",
                        font=("Segoe UI", 14, "bold"),
                        relief="flat",
                        bd=0,
                        padx=40,
                        pady=15,
                        cursor="hand2"
                    )
                    exit_btn.pack(side="bottom")
                    
                    # Status bar
                    status_label = tkinter.Label(
                        bottom_frame,
                        text=f"© 2025 CLASS VISION | Attendance completed for {Subject} • System Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        bg=DARK_BG,
                        fg=CYAN_COLOR,
                        font=("Segoe UI", 10)
                    )
                    status_label.pack(side="bottom", pady=(10, 0))
                    
                    # Auto-speak success message
                    text_to_speech(f"Attendance recorded successfully for {Subject}. Choose your next action.")
                    
                    options_window.mainloop()
                
                # Show the options dialog
                show_attendance_options()
                
            except:
                f = "No Face found for attendance"
                text_to_speech(f)
                cv2.destroyAllWindows()

    ###windo is frame for subject chooser
    # Modern UI Setup
    subject = Tk()
    subject.title("Choose Subject for Attendance")
    subject.geometry("650x450")
    subject.configure(bg="#e2e8f0")  # Matching background color
    subject.resizable(True, True)

    # Custom colors matching the main theme
    PRIMARY_COLOR = "#2563eb"
    SECONDARY_COLOR = "#7c3aed"
    ACCENT_COLOR = "#0ea5e9"
    BG_COLOR = "#e2e8f0"
    CARD_COLOR = "#e2e8f0"
    TEXT_COLOR = "#1e293b"
    SUCCESS_COLOR = "#16a34a"
    ERROR_COLOR = "#ef4444"

    # Main container
    main_frame = Frame(subject, bg=BG_COLOR)
    main_frame.pack(expand=True, fill=BOTH, padx=40, pady=30)

    # Header Section
    title = Label(main_frame, text="📚 Subject Selection", bg=BG_COLOR, fg=ACCENT_COLOR, font=("Segoe UI", 24, "bold"))
    title.pack(pady=(0, 10))
    
    subtitle = Label(main_frame, text="Choose subject for attendance tracking", bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 14))
    subtitle.pack(pady=(0, 30))

    # Input Section in Card
    input_card = Frame(main_frame, bg=CARD_COLOR, relief=FLAT, bd=0)
    input_card.pack(fill=X, pady=(0, 20))
    
    sub_label = Label(input_card, text="Enter Subject Name:", bg=CARD_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 14, "bold"))
    sub_label.pack(pady=(20, 10))
    
    tx = Entry(input_card, width=25, bg="#f1f5f9", fg=TEXT_COLOR, font=("Segoe UI", 16), relief=SOLID, bd=1, insertbackground=TEXT_COLOR)
    tx.pack(pady=(0, 20), ipady=10)
    
    # Notification area
    Notifica = Label(main_frame, text="", bg=BG_COLOR, fg=SUCCESS_COLOR, font=("Segoe UI", 12), wraplength=500)
    Notifica.pack(pady=(10, 20))

    # Modern Button Styles
    def modern_button(parent, text, command, color, emoji=""):
        btn = Button(parent, text=f"{emoji} {text}", command=command, bg=color, fg="white", 
                    font=("Segoe UI", 14, "bold"), relief=FLAT, bd=0, padx=30, pady=12, 
                    cursor="hand2", activebackground=color)
        btn.pack(pady=8, fill=X)
        
        # Hover effects
        def on_enter(e):
            btn.configure(relief=RAISED, bd=2)
        def on_leave(e):
            btn.configure(relief=FLAT, bd=0)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    def Attf():
        sub = tx.get()
        if sub == "":
            t = "Please enter the subject name!!!"
            text_to_speech(t)
            Notifica.configure(text="⚠️ Please enter a subject name!", fg=ERROR_COLOR)
        else:
            try:
                os.startfile(f"Attendance\\{sub}")
            except:
                Notifica.configure(text="❌ Subject folder not found!", fg=ERROR_COLOR)

    # Buttons Section
    buttons_frame = Frame(main_frame, bg=BG_COLOR)
    buttons_frame.pack(fill=X, pady=(20, 0))
    
    fill_a = modern_button(buttons_frame, "Fill Attendance", FillAttendance, PRIMARY_COLOR, "✅")
    attf = modern_button(buttons_frame, "Check Sheets", Attf, SECONDARY_COLOR, "📊")
    
    subject.mainloop()
